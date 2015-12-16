import cv2
import numpy
import openpyxl
from openpyxl.styles import PatternFill, Fill
import os

def main():
    img_file = os.getcwd() + '/zelda.jpg'
    img = cv2.imread(img_file)
    img = cv2.resize(img, (0,0), fx=0.5, fy=0.5)
    height, width, channels = img.shape

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Converted Image"

    for x in range(0, width):
        for y in range(0, height):
            bgr = img[y,x]
            bgr_hex = [hex(z).replace('0x', '') for z in bgr]
            bgr_hex.reverse()
            bgr_hex_padded = [''.join(('0',z)) if len(z) < 2 else z for z in bgr_hex]

            rgb_hex = ''.join(bgr_hex_padded)

            sheet.row_dimensions[y+1].height = .5
            sheet.column_dimensions[openpyxl.cell.get_column_letter(x+1)].width = .5
            sheet.cell(row=y+1, column=x+1).fill = PatternFill(fill_type='solid', start_color=rgb_hex, end_color=rgb_hex)

    wb.save(os.getcwd() + '/zelda.xlsx')

if __name__ == "__main__":
    main()